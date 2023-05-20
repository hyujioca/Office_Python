from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

ws_new = wb.create_sheet("追加シート")
value_data = 0
for x in range(1,11):
    for y in range(1,11):
        value_data = 10 * (x - 1) + y
        ws_new.cell(row=x, column=y).value = value_data
# Save the file
wb.save("sample.xlsx")


